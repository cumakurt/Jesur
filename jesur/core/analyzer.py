import os
import json
import re
import magic
import io
import docx
from threading import Lock
from typing import Optional
# import pdfplumber # Imported on demand to optional dependencies don't crash everything if missing
import openpyxl
from jesur.utils.cache import cache_manager

# Load patterns from JSON
PATTERNS_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'patterns.json')
PATTERNS = {}
try:
    with open(PATTERNS_FILE, 'r') as f:
        PATTERNS = json.load(f)
except FileNotFoundError:
    from jesur.utils.logger import log_warning
    log_warning(f"Patterns file not found: {PATTERNS_FILE}")
    PATTERNS = {"patterns": {}, "false_positives": [], "security_keywords": ""}
except json.JSONDecodeError as e:
    from jesur.utils.logger import log_error
    log_error(f"Invalid JSON in patterns file: {e}")
    PATTERNS = {"patterns": {}, "false_positives": [], "security_keywords": ""}
except Exception as e:
    from jesur.utils.logger import log_error
    log_error(f"Could not load patterns.json: {e}", exc_info=True)
    PATTERNS = {"patterns": {}, "false_positives": [], "security_keywords": ""}

# Extensions: treat as text for decoding when magic returns octet-stream or unknown.
# .txt is always a first-class target (credentials, vpn.txt, notes); keep it in this set.
# .tct included: occasional typo for .txt on SMB shares
TEXT_EXTENSIONS_FOR_CONTENT_SCAN = frozenset({
    '.txt', '.tct',
    '.cfg', '.ini', '.log', '.conf', '.env', '.properties', '.md', '.csv', '.xml',
    '.json', '.yml', '.yaml', '.vpn', '.toml', '.rc', '.sh', '.bat', '.ps1', '.cnf',
    '.config', '.ovpn', '.pem', '.crt', '.csr', '.key', '.cer', '.asc', '.pub',
    '.tf', '.tfvars', '.hcl', '.dockerfile', '.netrc', '.npmrc', '.pypirc', '.reg',
    '.plist', '.service', '.timer', '.sql', '.rdp',
})

# Built-in regex categories (merged with patterns.json; JSON keys override same name)
# Tuned to avoid minified JS (password:!0), Hunspell .dic (substring psk), Ruby Password::
# Separators: = : - – (not only =) for Notepad-style notes.
_PASSWORD_KEYWORDS = r'password|passwd|pwd'

DEFAULT_SENSITIVE_PATTERNS = {
    'password_assignment': (
        r'(?i)(?:' + _PASSWORD_KEYWORDS + r')(?!::)\s*[=:;\-–]\s*(?!true\b|false\b|!0\b|!1\b)([^\s;]{2,})'
    ),
    # Tab or multiple spaces between keyword and value (common in .txt / paste dumps)
    'password_spaced_value': (
        r'(?i)(?:' + _PASSWORD_KEYWORDS + r')(?:\t+|[\t ]{2,})(?!true\b|false\b|!0\b|!1\b)(\S{2,})'
    ),
    'secret_assignment': (
        r'(?i)\bsecret\b\s*[:=]\s*(?!true\b|false\b|!0\b|!1\b)([^\s;]{3,})'
    ),
    'api_key_assignment': (
        r'(?i)(api[_-]?key|access[_-]?token|client[_-]?secret)\s*[:=]\s*(?!true\b|false\b|!0\b|!1\b)(\S{8,})'
    ),
    'aws_access_key_id': r'\b(?:A3T[A-Z0-9]|AKIA|AGPA|AIDA|AROA|ASIA)[A-Z0-9]{16}\b',
    'aws_secret_access_key': (
        r'(?i)\baws[_-]?secret[_-]?access[_-]?key\b\s*[:=]\s*["\']?([A-Za-z0-9/+=]{40})'
    ),
    'google_api_key': r'\bAIza[0-9A-Za-z_-]{35}\b',
    'github_token': r'\b(?:ghp|gho|ghu|ghs|ghr)_[A-Za-z0-9_]{36,255}\b|\bgithub_pat_[A-Za-z0-9_]{80,255}\b',
    'gitlab_token': r'\bglpat-[A-Za-z0-9_-]{20,}\b',
    'slack_token': r'\bxox[baprs]-[A-Za-z0-9-]{10,}\b',
    'jwt_token': r'\beyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\b',
    'basic_auth_url': r'(?i)\b[a-z][a-z0-9+.-]{2,}://[^/\s:@]{2,}:[^@\s/]{2,}@[^\s\'"<>]+',
    'database_connection_string': (
        r'(?i)\b(?:jdbc:(?:mysql|postgresql|sqlserver|oracle)|mongodb(?:\+srv)?://|postgres(?:ql)?://|mysql://|mssql://|redis://)[^\s\'"<>]{8,}'
    ),
    'putty_private_key': r'(?im)^PuTTY-User-Key-File-\d+:',
    'vpn_or_tunnel': (
        r'(?i)(\bpsk\b|pre-shared|shared secret|auth-user-pass|remote\s+[\d.]+|tunnel\s+password)'
    ),
    'pem_private_block': r'-----BEGIN (RSA |OPENSSH |EC |DSA |ED25519 |)?PRIVATE KEY-----',
}

CRITICAL_PATTERN_CATEGORIES = frozenset({
    'aws_secret_access_key',
    'github_token',
    'gitlab_token',
    'slack_token',
    'putty_private_key',
    'pem_private_block',
    'database_connection_string',
})

HIGH_PATTERN_CATEGORIES = frozenset({
    'aws_access_key_id',
    'google_api_key',
    'jwt_token',
    'basic_auth_url',
    'password_assignment',
    'password_spaced_value',
    'secret_assignment',
    'api_key_assignment',
    'vpn_or_tunnel',
})


def _severity_for_category(category: str) -> str:
    """Map detection categories to report-friendly severities."""
    if category in CRITICAL_PATTERN_CATEGORIES:
        return 'Critical'
    if category in HIGH_PATTERN_CATEGORIES:
        return 'High'
    return 'Medium'


def _defaults_for_filename(filename: str) -> dict:
    """Drop noisy default categories for minified bundles and spell dictionaries."""
    fn = (filename or '').lower()
    out = dict(DEFAULT_SENSITIVE_PATTERNS)
    if fn.endswith(('.min.js', '.min.css', '.bundle.js')):
        for k in (
            'password_assignment', 'password_spaced_value',
            'secret_assignment', 'api_key_assignment', 'vpn_or_tunnel',
            'jwt_token', 'basic_auth_url',
        ):
            out.pop(k, None)
    elif fn.endswith(('.dic', '.aff')):
        out.pop('vpn_or_tunnel', None)
    return out


def _looks_like_text_bytes(data: bytes) -> bool:
    """Heuristic: mostly printable ASCII / common whitespace."""
    if not data:
        return False
    sample = data[:8192]
    if len(sample) < 4:
        return all(32 <= b <= 126 or b in (9, 10, 13) for b in sample)
    textish = sum(1 for b in sample if 32 <= b <= 126 or b in (9, 10, 13))
    return (textish / len(sample)) > 0.88


def _decode_text_bytes(file_content: bytes) -> Optional[str]:
    """
    Decode text files from SMB shares: UTF-8 (with/without BOM), UTF-16 LE/BE (Windows Notepad),
    then Latin-1 fallback.

    UTF-16 LE without BOM must be tried BEFORE utf-8: otherwise Python accepts UTF-16-ASCII bytes
    as valid UTF-8 (NUL U+0000 between each Latin letter), and regex never matches real words.
    """
    if not file_content:
        return None
    # UTF-8 BOM
    if file_content.startswith(b'\xef\xbb\xbf'):
        try:
            return file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            pass
    # UTF-16 LE / BE with BOM
    if file_content.startswith(b'\xff\xfe'):
        try:
            return file_content.decode('utf-16-le')
        except UnicodeDecodeError:
            try:
                return file_content.decode('utf-16')
            except UnicodeDecodeError:
                pass
    if file_content.startswith(b'\xfe\xff'):
        try:
            return file_content.decode('utf-16-be')
        except UnicodeDecodeError:
            pass
    # UTF-16 without BOM (common on Windows) — MUST run before utf-8 (see docstring)
    if len(file_content) >= 4 and len(file_content) % 2 == 0:
        nul_ratio = file_content.count(0) / len(file_content)
        if nul_ratio > 0.15:
            try:
                return file_content.decode('utf-16-le')
            except UnicodeDecodeError:
                try:
                    return file_content.decode('utf-16-be')
                except UnicodeDecodeError:
                    pass
    try:
        return file_content.decode('utf-8')
    except UnicodeDecodeError:
        pass
    try:
        return file_content.decode('utf-8-sig')
    except UnicodeDecodeError:
        pass
    try:
        return file_content.decode('latin-1')
    except UnicodeDecodeError:
        return None


def _extract_text_for_pattern_scan(
    file_content: bytes, file_type: str, filename: str
) -> Optional[str]:
    """Extract searchable text for pattern matching (MIME + extension fallbacks)."""
    fn = (filename or '').lower()

    if file_type.startswith('application/pdf'):
        return extract_pdf_content(file_content)
    if file_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
        return extract_docx_content(file_content)
    if file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        return extract_xlsx_content(file_content)
    if file_type.startswith('text/'):
        return _decode_text_bytes(file_content)

    _, ext = os.path.splitext(fn)
    if ext in TEXT_EXTENSIONS_FOR_CONTENT_SCAN:
        return _decode_text_bytes(file_content)
    if _looks_like_text_bytes(file_content):
        return _decode_text_bytes(file_content)
    return None


# Global magic instance for better performance
_magic_instance = None
_magic_lock = Lock()

def _get_magic_instance():
    """Get or create global magic instance."""
    global _magic_instance
    with _magic_lock:
        if _magic_instance is None:
            _magic_instance = magic.Magic(mime=True)
        return _magic_instance

def get_file_type(file_content: bytes) -> str:
    """
    Detect file type using magic library.
    
    Args:
        file_content: File content as bytes
        
    Returns:
        MIME type string or "unknown" if detection fails
    """
    try:
        mime = _get_magic_instance()
        with _magic_lock:
            file_type = mime.from_buffer(file_content)
        
        # Excel and Word checks
        if file_type == 'application/vnd.ms-excel' or \
           file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or \
           file_type.startswith('application/vnd.ms-excel.') or \
           '.xls' in file_type.lower():
            return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        if file_type == 'application/msword' or \
           file_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' or \
           file_type.startswith('application/vnd.ms-word.') or \
           '.doc' in file_type.lower():
            return 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            
        return file_type
    except (IOError, OSError) as e:
        from jesur.utils.logger import log_debug
        log_debug(f"File type detection IO error: {e}")
        return "unknown"
    except Exception as e:
        from jesur.utils.logger import log_debug
        log_debug(f"File type detection failed: {e}")
        return "unknown"

def extract_pdf_content(file_content: bytes) -> str:
    """
    Extract text content from PDF file.
    
    Args:
        file_content: PDF file content as bytes
        
    Returns:
        Extracted text content or None if extraction fails
    """
    try:
        import pdfplumber
        with io.BytesIO(file_content) as pdf_file:
            with pdfplumber.open(pdf_file) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
        return text
    except ImportError:
        from jesur.utils.logger import log_debug
        log_debug("pdfplumber not available, PDF extraction skipped")
        return None
    except (IOError, OSError) as e:
        from jesur.utils.logger import log_debug
        log_debug(f"PDF extraction IO error: {e}")
        return None
    except Exception as e:
        from jesur.utils.logger import log_debug
        log_debug(f"PDF extraction failed: {e}")
        return None

def extract_docx_content(file_content: bytes) -> str:
    """
    Extract text content from DOCX file.
    
    Args:
        file_content: DOCX file content as bytes
        
    Returns:
        Extracted text content or None if extraction fails
    """
    try:
        doc = docx.Document(io.BytesIO(file_content))
        return " ".join([paragraph.text for paragraph in doc.paragraphs])
    except (IOError, OSError) as e:
        from jesur.utils.logger import log_debug
        log_debug(f"DOCX extraction IO error: {e}")
        return None
    except Exception as e:
        from jesur.utils.logger import log_debug
        log_debug(f"DOCX extraction failed: {e}")
        return None

def extract_xlsx_content(file_content: bytes) -> str:
    """
    Extract text content from XLSX file.
    
    Args:
        file_content: XLSX file content as bytes
        
    Returns:
        Extracted text content or None if extraction fails
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                text.extend(str(cell.value) for cell in row if cell.value is not None)
        return " ".join(text)
    except (IOError, OSError) as e:
        from jesur.utils.logger import log_debug
        log_debug(f"XLSX extraction IO error: {e}")
        return None
    except Exception as e:
        from jesur.utils.logger import log_debug
        log_debug(f"XLSX extraction failed: {e}")
        return None

def check_sensitive_patterns(
    file_content: bytes, file_type: str, ip: str, filename: str = ''
) -> list:
    """
    Check file content for sensitive patterns (defaults + patterns.json).

    JSON overrides categories with the same key name. When patterns.json was
    empty, no content matches were found; defaults fix common cases (password=,
    VPN secrets, PEM blocks) and text extraction for .txt even if magic says
    application/octet-stream.
    """
    results_list = []
    content = _extract_text_for_pattern_scan(file_content, file_type, filename)

    if not content:
        return results_list

    # Check security keywords (optional regex in JSON)
    sec_keywords = PATTERNS.get('security_keywords', '')
    if sec_keywords:
        security_pattern = cache_manager.get_compiled_pattern(sec_keywords)
        if security_pattern.search(content):
            results_list.append({
                'category': 'security_keyword',
                'match': 'Contains security-related keywords',
                'severity': 'Medium',
            })

    user_patterns = PATTERNS.get('patterns') or {}
    sensitive_patterns = {**_defaults_for_filename(filename), **user_patterns}
    false_positives = PATTERNS.get('false_positives', [])
    compiled_fps = [cache_manager.get_compiled_pattern(fp) for fp in false_positives]

    from jesur.core.constants import MAX_MATCHES_PER_CATEGORY
    category_counts = {}

    lines = content.split('\n')
    for category, pattern in sensitive_patterns.items():
        compiled_pattern = cache_manager.get_compiled_pattern(pattern)
        category_counts[category] = 0

        for line in lines:
            if category_counts[category] >= MAX_MATCHES_PER_CATEGORY:
                break

            for match in compiled_pattern.finditer(line):
                if category_counts[category] >= MAX_MATCHES_PER_CATEGORY:
                    break

                match_str = match.group(0)
                if not any(fp.search(match_str) for fp in compiled_fps):
                    results_list.append({
                        'category': category,
                        'match': match_str,
                        'severity': _severity_for_category(category),
                    })
                    category_counts[category] += 1

    return results_list
